from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import pyodbc
import pandas as pd
from openpyxl import load_workbook, workbook, worksheet
##---------------------------------------------------------------------------------------------------------------------

def obter_conexao():
    server = 'NTB-KAIQUE\SQLEXPRESS'
    database = 'REL_DIZIMISTA'
    username = 'sa'
    password = '123'
    driver = 'SQL Server Native Client 11.0'


    conexao_string = f'DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}'
    return pyodbc.connect(conexao_string)


##--------------------------------------------------------------------------------------------------------------------
#pagina principal

app = Flask(__name__)
app.config['SECRET_KEY'] = 'TESTE'

@app.route('/')
def inicio():
    return render_template('index.html')

##--------------------------------------------------------------------------------------------------------------------
##visualizar dizimista

@app.route('/vis_dizimista', methods=['POST','GET'])
def vis_dizimista():
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM DIZIMISTA")
        dados = cursor.fetchall()

    except Exception as e:
        mensagem = "erro ao visualizar registro"
        return mensagem

    finally:
        cursor.close()
        conn.close() 

    return render_template ('vis_dizimista.html', dados=dados)


##--------------------------------------------------------------------------------------------------------------------
#ins_dizimista
@app.route('/ins_dizimista', methods=['POST'])
def ins_diz():
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM DIZIMISTA")
        dados = cursor.fetchall()

    except Exception as e:
        mensagem = "erro ao visualizar registro"
        return mensagem

    finally:
        cursor.close()
        conn.close() 

    return render_template ('ins_dizimista.html', dados=dados)

##--------------------------------------------------------------------------------------------------------------------

@app.route('/inserir', methods=['POST'])
def inserir_dados():
    cargo = request.form['cargo']
    nome = request.form['nome']
    id = request.form['id']

    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        cursor.execute("INSERT INTO DIZIMISTA (CARGO,NOME,RI) VALUES (?,?,?)", cargo,nome,id)
        conn.commit()
        flash("Registro inserido com sucesso!")
        return redirect(url_for('vis_dizimista'))
    
    except Exception as e:
        conn.rollback()
        mensagem =f"Erro ao inserir o registro: {str(e)}"

    finally:
        cursor.close()
        conn.close()

    return mensagem

##--------------------------------------------------------------------------------------------------------------------
#excluir dizimista

@app.route('/exc_dizimista', methods=['POST'])
def exc_diz():
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM DIZIMISTA")
        dados = cursor.fetchall()

    except Exception as e:
        mensagem = "erro ao visualizar registro"
        return mensagem

    finally:
        cursor.close()
        conn.close() 

    return render_template ('exc_dizimista.html', dados=dados)

##================================================================
@app.route('/exclusao', methods=['POST'])

def exc_dizimista():
    exclusao = request.form['exclusao']
    conn = obter_conexao()
    cursor = conn.cursor()
    dizimista = 'DIZIMISTA'

    try:
        comando =f'DELETE FROM {dizimista} WHERE ID = ? OR RI = ?'
        cursor.execute(comando,(exclusao,exclusao))
        conn.commit()
        flash("Registro excluído com sucesso!")
        return redirect(url_for('vis_dizimista'))
    

    except Exception as e:
        mensagem = f"Erro ao excluir o registro: {str(e)}"        

    finally:
        cursor.close()
        conn.close() 

    return mensagem

##--------------------------------------------------------------------------------------------------------------------
@app.route('/alt_dizimista', methods=['POST'])
def alt_diz():
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM DIZIMISTA")
        dados = cursor.fetchall()

    except Exception as e:
        mensagem = "erro ao visualizar registro"
        return mensagem

    finally:
        cursor.close()
        conn.close() 

    return render_template ('alt_dizimista.html', dados=dados)

#==============================================================
@app.route('/alteracao', methods=['POST'])
def alteracao():
    id_ou_ri = request.form['id_ou_ri']
    campo = request.form['campo']
    alteracao = request.form['alteracao']


    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        comando = f'UPDATE DIZIMISTA SET {campo} = ? WHERE ID = ? OR RI = ?'
        cursor.execute(comando,(alteracao,id_ou_ri,id_ou_ri))
        conn.commit()
        flash("Registro atualizado com sucesso!")
        return redirect(url_for('vis_dizimista'))
    

    except Exception as e:
        mensagem = f"Erro ao atualizar o registro: {str(e)}"        

    finally:
        cursor.close()
        conn.close() 

    return mensagem

##--------------------------------------------------------------------------------------------------------------------
#visualizar movimento (contribuição)

@app.route('/vis_movimento', methods=['POST','GET'])
def vis_movimento():
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT A.ID, A.VALOR, A.DATA, A.ID_DIZIMISTA, B.NOME FROM CONTRIBUICAO A INNER JOIN DIZIMISTA B ON B.ID = A.ID_DIZIMISTA ")
        dados = cursor.fetchall()

    except Exception as e:
        mensagem = "erro ao visualizar registro"
        return mensagem

    finally:
        cursor.close()
        conn.close() 
    return render_template ('vis_movimento.html', dados=dados)
##--------------------------------------------------------------------------------------------------------------------
#inserir movimento (contribuição)

@app.route('/ins_movimento', methods=['POST','GET'])
def ins_movimento():
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM DIZIMISTA")
        dados = cursor.fetchall()

    except Exception as e:
        mensagem = "erro ao visualizar registro"
        return mensagem

    finally:
        cursor.close()
        conn.close() 
    return render_template ('ins_movimento.html', dados=dados)
##--------------------------------------------------------------------------------------------------------------------
@app.route('/ins_mov_diz', methods=['POST'])
def ins_mov_diz():
    id_dizimista = request.form['id_dizimista']
    data = request.form['data']
    valor = request.form['valor']

    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        comando = f'INSERT INTO CONTRIBUICAO (ID_DIZIMISTA,DATA,VALOR) VALUES (?,?,?)'
        cursor.execute(comando,id_dizimista,data,valor)
        conn.commit()
        flash("Registro inserido com sucesso!")
        return redirect(url_for('ins_movimento'))
    

    except Exception as e:
        mensagem = f"Erro ao atualizar o registro: {str(e)}"        

    finally:
        cursor.close()
        conn.close() 

    return mensagem

##--------------------------------------------------------------------------------------------------------------------
#Alteração de movimento
@app.route('/alt_movimento', methods=['POST','GET'])
def alt_movimento():
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT A.ID, A.VALOR, A.DATA, A.ID_DIZIMISTA, B.NOME FROM CONTRIBUICAO A INNER JOIN DIZIMISTA B ON B.ID = A.ID_DIZIMISTA ")
        dados = cursor.fetchall()

    except Exception as e:
        mensagem = f"erro ao visualizar registro: {str(e)}"  
        return mensagem

    finally:
        cursor.close()
        conn.close() 

    return render_template ('alt_movimento.html', dados=dados)
##--------------------------------------------------------------------------------------------------------------------

#Filtrando por mês na alteração de movimento

@app.route('/alteracao_mov', methods=['POST'])
def alteracao_mov():

    mes_movimento = request.form['mes_movimento']
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        consulta = f"SELECT A.ID, A.VALOR, A.DATA, A.ID_DIZIMISTA, B.NOME FROM CONTRIBUICAO A INNER JOIN DIZIMISTA B ON B.ID = A.ID_DIZIMISTA WHERE A.DATA LIKE ? "
        cursor.execute(consulta,f'%{mes_movimento}%')
        dados = cursor.fetchall()

    except Exception as e:
        mensagem = f"erro ao visualizar registro: {str(e)}"  
        return mensagem

    finally:
        cursor.close()
        conn.close() 

    return render_template ('alt_movimento2.html', dados=dados)

##--------------------------------------------------------------------------------------------------------------------

@app.route('/alteracao_mov2', methods=['POST'])
def alteracao_mov2():

    id_movimento = request.form['id_movimento']
    campo_mov = request.form['campo_mov']
    atual_campo_mov = request.form['atual_campo_mov']
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        consulta = f"UPDATE CONTRIBUICAO SET {campo_mov} = ? WHERE ID = ?"
        cursor.execute(consulta,atual_campo_mov,id_movimento)
        conn.commit()
        flash("Registro atualizado com sucesso!")
        return redirect(url_for('vis_movimento'))
    
    except Exception as e:
        mensagem = f"erro ao visualizar registro: {str(e)}"  

    finally:
        cursor.close()
        conn.close() 

    return mensagem
##--------------------------------------------------------------------------------------------------------------------
##acessar rotima para excluir movimento

@app.route('/exc_movimento', methods=['POST','GET'])
def exc_movimento():
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT A.ID, A.VALOR, A.DATA, A.ID_DIZIMISTA, B.NOME FROM CONTRIBUICAO A INNER JOIN DIZIMISTA B ON B.ID = A.ID_DIZIMISTA ")
        dados = cursor.fetchall()

    except Exception as e:
        mensagem = f"erro ao visualizar registro: {str(e)}"  
        return mensagem

    finally:
        cursor.close()
        conn.close() 

    return render_template ('exc_movimento.html', dados=dados)

##--------------------------------------------------------------------------------------------------------------------
## filtrar para excluir movimento

@app.route('/exclusao_mov', methods=['POST','GET'])
def exclusao_mov():
   
    mes_exclusao = request.form['mes_exclusao']
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        consulta = f"SELECT A.ID, A.VALOR, A.DATA, A.ID_DIZIMISTA, B.NOME FROM CONTRIBUICAO A INNER JOIN DIZIMISTA B ON B.ID = A.ID_DIZIMISTA WHERE A.DATA LIKE ? "
        cursor.execute(consulta,f'%{mes_exclusao}%')
        dados = cursor.fetchall()

    except Exception as e:
        mensagem = f"erro ao visualizar registro: {str(e)}"  
        return mensagem

    finally:
        cursor.close()
        conn.close() 

    return render_template ('exc_movimento2.html', dados=dados)

##--------------------------------------------------------------------------------------------------------------------
## excluir movimento

@app.route('/exclusao_mov2', methods=['POST','GET'])
def exclusao_mov2():
   
    exclusao_id_mov = request.form['exclusao_id_mov']
        
    conn = obter_conexao()
    cursor = conn.cursor()

    try:
        consulta = f"DELETE FROM CONTRIBUICAO WHERE ID = ? "
        cursor.execute(consulta,exclusao_id_mov)
        conn.commit()
        flash("Registro excluído com sucesso!")
        return redirect(url_for('vis_movimento'))
    
    except Exception as e:
        mensagem = f"Erro ao excluir o registro: {str(e)}"        

    finally:
        cursor.close()
        conn.close() 

    return mensagem



##--------------------------------------------------------------------------------------------------------------------

#Geração do relatório
@app.route('/gerar_relatorio', methods=['POST'])
def gerar_relatorio():
    return render_template('gerar_relatorio.html')

##--------------------------------------------------------------------------------------------------------------------

# Caminho do relatório
@app.route('/relatorio', methods=['POST', 'GET'])
def relatorio():
    mes_para_relatorio = request.form['mes_para_relatorio']
    ano_para_relatorio = request.form['ano_para_relatorio']
    tipo_relatorio = request.form['tipo_relatorio']

    conn = obter_conexao()
    cursor = conn.cursor()

    if tipo_relatorio == 'SINTÉTICO':
        consulta = f"SELECT DISTINCT NOME, CARGO, RI FROM CONTRIBUICAO A INNER JOIN DIZIMISTA B ON A.ID_DIZIMISTA = B.ID WHERE A.DATA LIKE ? AND A.DATA LIKE ? ORDER BY CARGO DESC"
        dados = pd.read_sql(consulta, conn, params=[f'%{mes_para_relatorio}%', f'{ano_para_relatorio}%'])
        cursor.execute(consulta, [f'%{mes_para_relatorio}%', f'{ano_para_relatorio}%'])

        resultado = cursor.fetchall()
        count = 0


        if mes_para_relatorio == '01':
            nome_mes = 'JANEIRO'

        elif mes_para_relatorio == '02':
            nome_mes = 'FEVEREIRO'

        elif mes_para_relatorio == '03':
            nome_mes = 'MARÇO'

        elif mes_para_relatorio == '04':
            nome_mes = 'ABRIL'

        elif mes_para_relatorio == '05':
            nome_mes = 'MAIO'

        elif mes_para_relatorio == '06':
            nome_mes = 'JUNHO'

        elif mes_para_relatorio == '07':
            nome_mes = 'JULHO'

        elif mes_para_relatorio == '08':
            nome_mes = 'AGOSTO'

        elif mes_para_relatorio == '09':
            nome_mes = 'SETEMBRO'

        elif mes_para_relatorio == '10':
            nome_mes = 'OUTUBRO'

        elif mes_para_relatorio == '11':
            nome_mes = 'NOVEMBRO'

        elif mes_para_relatorio == '12':
            nome_mes = 'DEZEMBRO'

        excel = "C:/RELATORIO DIZIMISTA/relatorio_sintetico.ods"
        pag_excel = "aaa"

        relatorio = load_workbook(excel)
        pagina = relatorio[pag_excel]

        wb = relatorio
        planilha = wb.worksheets[0]

        planilha['A5'] = f"DIZIMISTAS REFERENTE {nome_mes} - {ano_para_relatorio}"


        for i in resultado:
            planilha['A9'] = (resultado[0][1])
            planilha['B9'] =  (resultado[0][0])
            planilha['I9'] =  (resultado[0][2])

            planilha['A10'] =  (resultado[1][1])
            planilha['B10'] =  (resultado[1][0])
            planilha['I10'] =  (resultado[1][2])

            planilha['A11'] =  (resultado[2][1])
            planilha['B11'] =  (resultado[2][0])
            planilha['I11'] =  (resultado[2][2])

            planilha['A12'] =  (resultado[3][1])
            planilha['B12'] =  (resultado[3][0])
            planilha['I12'] =  (resultado[3][2])

            planilha['A13'] =  (resultado[4][1])
            planilha['B13'] =  (resultado[4][0])
            planilha['I13'] =  (resultado[4][2])

            planilha['A14'] =  (resultado[5][1])
            planilha['B14'] =  (resultado[5][0])
            planilha['I14'] =  (resultado[5][2])

            planilha['A15'] =  (resultado[6][1])
            planilha['B15'] =  (resultado[6][0])
            planilha['I15'] =  (resultado[6][2])

            planilha['A16'] =  (resultado[7][1])
            planilha['B16'] =  (resultado[7][0])
            planilha['I16'] =  (resultado[7][2])

            planilha['A17'] =  (resultado[8][1])
            planilha['B17'] =  (resultado[8][0])
            planilha['I17'] =  (resultado[8][2])

            planilha['A18'] =  (resultado[9][1])
            planilha['B18'] =  (resultado[9][0])
            planilha['I18'] =  (resultado[9][2])
            
            planilha['A19'] =  (resultado[10][1])
            planilha['B19'] =  (resultado[10][0])
            planilha['I19'] =  (resultado[10][2])

    ## aqui ele gera o arquivo
        # arquivo_gerado = f"{nome_mes}_{ano_para_relatorio}.xlsx"
        # dados.to_excel(arquivo_gerado, index=False)

        relatorio.save(excel)

        pegar_conteudo = pd.read_excel("relatorio_sintetico.xlsx")
        arquivo_gerado = f"{nome_mes}_{ano_para_relatorio}.xlsx"
        pegar_conteudo.to_excel(arquivo_gerado, index=False)

        # Fechar a conexão com o banco de dados
        conn.close()

        flash(f'Dados exportados para {excel}')
        return redirect(url_for('vis_movimento'))

    
    elif tipo_relatorio == 'ANALÍTICO':
        consulta1 = f"SELECT CONVERT(varchar(10),DATA,103) AS DATA,CARGO,NOME,VALOR, RI FROM CONTRIBUICAO A INNER JOIN DIZIMISTA B ON A.ID_DIZIMISTA = B.ID WHERE A.DATA LIKE ? AND A.DATA LIKE ? ORDER BY DATA "
        dados = pd.read_sql(consulta1, conn, params=[f'%{mes_para_relatorio}%', f'{ano_para_relatorio}%'])
        cursor.execute(consulta1, [f'%{mes_para_relatorio}%', f'{ano_para_relatorio}%'])

        resultado1 = cursor.fetchall()

        consulta2 = "SELECT SUM(VALOR) FROM CONTRIBUICAO A INNER JOIN DIZIMISTA B ON A.ID_DIZIMISTA = B.ID WHERE A.DATA LIKE ? AND A.DATA LIKE ? "
        dados2 = pd.read_sql(consulta2, conn, params=[f'%{mes_para_relatorio}%', f'{ano_para_relatorio}%'])
        cursor.execute(consulta2, [f'%{mes_para_relatorio}%', f'{ano_para_relatorio}%'])

        soma_valor = cursor.fetchall()


        if mes_para_relatorio == '01':
            nome_mes = 'JANEIRO'

        elif mes_para_relatorio == '02':
            nome_mes = 'FEVEREIRO'

        elif mes_para_relatorio == '03':
            nome_mes = 'MARÇO'

        elif mes_para_relatorio == '04':
            nome_mes = 'ABRIL'

        elif mes_para_relatorio == '05':
            nome_mes = 'MAIO'

        elif mes_para_relatorio == '06':
            nome_mes = 'JUNHO'

        elif mes_para_relatorio == '07':
            nome_mes = 'JULHO'

        elif mes_para_relatorio == '08':
            nome_mes = 'AGOSTO'

        elif mes_para_relatorio == '09':
            nome_mes = 'SETEMBRO'

        elif mes_para_relatorio == '10':
            nome_mes = 'OUTUBRO'

        elif mes_para_relatorio == '11':
            nome_mes = 'NOVEMBRO'

        elif mes_para_relatorio == '12':
            nome_mes = 'DEZEMBRO'

        excel = "C:/RELATORIO DIZIMISTA/relatorio_analitico.xlsx"
        pag_excel = "aaa"

        relatorio1 = load_workbook(excel)
        pagina = relatorio1[pag_excel]

        wb = relatorio1
        planilha = wb.worksheets[0]

        planilha['A5'] = f"DIZIMISTAS REFERENTE {nome_mes} - {ano_para_relatorio}"


        for i in resultado1:
            planilha['A9'] =  (resultado1[0][0])
            planilha['B9'] =  (resultado1[0][1])
            planilha['C9'] =  (resultado1[0][2])
            planilha['G9'] =  (resultado1[0][3])
            planilha['I9'] =  (resultado1[0][4])


            planilha['A10'] =  (resultado1[1][0])
            planilha['B10'] =  (resultado1[1][1])
            planilha['C10'] =  (resultado1[1][2])
            planilha['G10'] =  (resultado1[1][3])
            planilha['I10'] =  (resultado1[1][4])

            planilha['A11'] =  (resultado1[2][0])
            planilha['B11'] =  (resultado1[2][1])
            planilha['C11'] =  (resultado1[2][2])
            planilha['G11'] =  (resultado1[2][3])
            planilha['I11'] =  (resultado1[2][4])

            planilha['A12'] =  (resultado1[3][0])
            planilha['B12'] =  (resultado1[3][1])
            planilha['C12'] =  (resultado1[3][2])
            planilha['G12'] =  (resultado1[3][3])
            planilha['I12'] =  (resultado1[3][4])

            planilha['A13'] =  (resultado1[4][0])
            planilha['B13'] =  (resultado1[4][1])
            planilha['C13'] =  (resultado1[4][2])
            planilha['G13'] =  (resultado1[4][3])
            planilha['I13'] =  (resultado1[4][4])

            planilha['A14'] =  (resultado1[5][0])
            planilha['B14'] =  (resultado1[5][1])
            planilha['C14'] =  (resultado1[5][2])
            planilha['G14'] =  (resultado1[5][3])
            planilha['I14'] =  (resultado1[5][4])

            planilha['A15'] =  (resultado1[6][0])
            planilha['B15'] =  (resultado1[6][1])
            planilha['C15'] =  (resultado1[6][2])
            planilha['G15'] =  (resultado1[6][3])
            planilha['I15'] =  (resultado1[6][4])

            planilha['A16'] =  (resultado1[7][0])
            planilha['B16'] =  (resultado1[7][1])
            planilha['C16'] =  (resultado1[7][2])
            planilha['G16'] =  (resultado1[7][3])
            planilha['I16'] =  (resultado1[7][4])

            planilha['A17'] =  (resultado1[8][0])
            planilha['B17'] =  (resultado1[8][1])
            planilha['C17'] =  (resultado1[8][2])
            planilha['G17'] =  (resultado1[8][3])
            planilha['I17'] =  (resultado1[8][4])

            planilha['A18'] =  (resultado1[9][0])
            planilha['B18'] =  (resultado1[9][1])
            planilha['C18'] =  (resultado1[9][2])
            planilha['G18'] =  (resultado1[9][3])
            planilha['I18'] =  (resultado1[9][4])

            planilha['A19'] =  (resultado1[10][0])
            planilha['B19'] =  (resultado1[10][1])
            planilha['C19'] =  (resultado1[10][2])
            planilha['G19'] =  (resultado1[10][3])
            planilha['I19'] =  (resultado1[10][4])

            planilha['A20'] =  (resultado1[11][0])
            planilha['B20'] =  (resultado1[11][1])
            planilha['C20'] =  (resultado1[11][2])
            planilha['G20'] =  (resultado1[11][3])
            planilha['I20'] =  (resultado1[11][4])

            planilha['A21'] =  (resultado1[12][0])
            planilha['B21'] =  (resultado1[12][1])
            planilha['C21'] =  (resultado1[12][2])
            planilha['G21'] =  (resultado1[12][3])
            planilha['I21'] =  (resultado1[12][4])


            planilha['G27'] = (soma_valor[0][0])

    ## aqui ele gera o arquivo
        # arquivo_gerado = f"{nome_mes}_{ano_para_relatorio}.xlsx"
        # dados.to_excel(arquivo_gerado, index=False)

        relatorio1.save(excel)
        # Fechar a conexão com o banco de dados
        conn.close()

        flash(f'Dados exportados para {excel}')
        return redirect(url_for('inicio'))


if __name__ == '__main__':
    app.run(debug=True)


