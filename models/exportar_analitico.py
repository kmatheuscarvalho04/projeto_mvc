from projeto_mvc.models.db import obter_conexao
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os

def gerar_excel_analitico(dados, mes_relatorio):
    # Transforma '202505' -> 'MAIO - 2025'
    ano = mes_relatorio[:4]
    mes = mes_relatorio[4:]
    nome_mes = calendar.month_name[int(mes)].upper()
    nome_aba = f"{nome_mes} - {ano}"

    # Cria a planilha
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba

    # Cabeçalhos na linha 1 (ajuste conforme necessário)
    cabecalhos = ["CARGO", "NOME", "RI"]
    for col, cabecalho in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col, value=cabecalho)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Dados a partir da linha 10
    for i, linha in enumerate(dados, start=10):
        for j, valor in enumerate(linha, start=1):
            ws.cell(row=i, column=j, value=valor)

    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Garante que a pasta existe
    os.makedirs("relatorios_gerados", exist_ok=True)

    # Nome do arquivo
    nome_arquivo = f"Relatorio_Analitico_{nome_mes}_{ano}.xlsx"
    caminho = f"relatorios_gerados/{nome_arquivo}"

    # Salva o arquivo
    wb.save(caminho)
    return caminho
